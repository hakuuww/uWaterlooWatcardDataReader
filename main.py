import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from columnar import columnar
from click import style
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

#################################################
#Enter you uWaterloo student number and password here (replacing the xxx)
#Enter your desired starting and end date for transaction data in the format 'mm/dd/year'
#The default start date is '01/01/2022', you can leave the trans_end_date empty if you want,
#it will just end at the current date
student_number = '20935781'
password = 'Yushan100!'
trans_start_date = '01/01/2021'
trans_end_date = ''
#################################################

driver = webdriver.Chrome(executable_path='C:/SeleniumDrivers/chromedriver.exe')

driver.get("https://watcard.uwaterloo.ca/OneWeb/Account/LogOn") #open the link in the browser

driver.implicitly_wait(30)

########################################Enter passwords and click log on on the Watcard login page######################
account_element = driver.find_element(By.ID, "Account")
password_element = driver.find_element(By.ID, "Password")
submit_element = driver.find_element(By.CLASS_NAME, "btn")
account_element.click()
account_element.send_keys(student_number)
password_element.click()
password_element.send_keys(password) 
submit_element.click()
########################################################################################################################

############################BALANCES################################
financial_toggle_element = driver.find_element(By.LINK_TEXT, "FINANCIAL")
financial_toggle_element.click()
balances_element = driver.find_element(By.LINK_TEXT, "BALANCES")
balances_element.click()
# /html/body/div[1]/div[2]/section/div[2]/div[1]/div/table/tbody/tr[1]/td[1]


driver.implicitly_wait(100)
balance_rows = len(driver.find_elements(By.XPATH, "/html/body/div[1]/div[2]/section/div[2]/div[1]/div/table/tbody/tr"))
balance_columns = len(driver.find_elements(By.XPATH, "/html/body/div[1]/div[2]/section/div[2]/div[1]/div/table/tbody/tr[1]/td"))
print(f"Watcard Balances:")
print(f"({balance_rows} rows and {balance_columns} columns in the table)")

#/html/body/div[1]/div[2]/section/div[2]/div[1]/div/table/thead/tr/th[1]
balance_header_list=[]
for h in range(1,balance_columns+1):
    balance_header_col_temp = str(h)
    balance_header_data = driver.find_element(By.XPATH,"/html/body/div[1]/div[2]/section/div[2]/div[1]/div/table/thead/tr/th["+balance_header_col_temp+"]").text
    balance_header_list.append(balance_header_data)
#print(balance_header_list)

# /html/body/div[1]/div[2]/section/div[2]/div[1]/div/table/tbody/tr[2]/td[1]
balance_final_data = []
for r in range(1,balance_rows+1):
    list_temp = []
    for c in range (1,balance_columns+1):
        balance_row_temp = str(r)
        balance_col_temp = str(c)
        balance_data = driver.find_element(By.XPATH,"/html/body/div[1]/div[2]/section/div[2]/div[1]/div/table/tbody/tr["+balance_row_temp+"]/td["+balance_col_temp+"]").text
        print(balance_data, '      ', end ='')
        list_temp.append(balance_data)
    print()
    #print('{0:<25} {1:>10} {2:>4} {3:>4} {4:>30} {5:>30}'.format(*list_temp))
    balance_final_data.append(list_temp)

balance_patterns = [
    ('#', lambda text: style(text, fg='blue')),
    ('Name', lambda text: style(text, fg='cyan')),
    ('Credit', lambda text: style(text, fg='green')),
    ('Amount', lambda text: style(text, fg='white')),
    ('RESIDENCE', lambda text: style(text, fg='white' , bg='green')),
    ('FLEXIBLE', lambda text: style(text, fg='white' , bg='green')),
    ]

balance_table = columnar(balance_final_data, headers=balance_header_list, patterns=balance_patterns, wrap_max=0,max_column_width=30,terminal_width=200)
######################################################################################


##########################TRANSACTIONS########################################
driver.back()
financial_toggle_element = driver.find_element(By.LINK_TEXT, "FINANCIAL")
financial_toggle_element.click()
transactions_element = driver.find_element(By.LINK_TEXT, "TRANSACTIONS")
transactions_element.click()

driver.implicitly_wait(30)
trans_start_date_element = driver.find_element(By.ID, "trans_start_date")
trans_end_date_element = driver.find_element(By.ID, "trans_end_date")
trans_end_date_value_string = trans_end_date_element.get_attribute('value')

trans_start_date_element.click()
trans_start_date_element.clear()
trans_start_date_element.send_keys(trans_start_date)



if trans_end_date != '' :
    trans_end_date_element.click()
    trans_end_date_element.clear()
    trans_end_date_element.send_keys(trans_end_date)
    trans_end_date_value_string = trans_end_date_element



trans_search_element = driver.find_element(By.ID, "trans_search")
trans_search_element.click()


driver.implicitly_wait(100)
rows = len(driver.find_elements(By.XPATH, "/html/body/div[1]/div[2]/section/div[2]/div/div/div[2]/table/tbody/tr"))
columns = len(driver.find_elements(By.XPATH, "/html/body/div[1]/div[2]/section/div[2]/div/div/div[2]/table/tbody/tr[1]/td"))
print(f"Watcard Transactions starting from: {trans_start_date} to {trans_end_date_value_string}")
print(f"{rows-1} transactions in total")
print(f"({rows} rows and {columns} columns in the table)")
print('Loading transaction data table')
#/html/body/div[1]/div[2]/section/div[2]/div/div/div[2]/table/tbody
#/html/body/div[1]/div[2]/section/div[2]/div/div/div[2]/table/tbody/tr[1]/td[1]
#xpath of headers:          /html/body/div[1]/div[2]/section/div[2]/div/div/div[2]/table/thead/tr/th[1]

driver.implicitly_wait(100)
header_list=[]
for h in range(1,columns+1):
    header_col_temp = str(h)
    header_data = driver.find_element(By.XPATH,"/html/body/div[1]/div[2]/section/div[2]/div/div/div[2]/table/thead/tr/th["+header_col_temp+"]").text
    header_list.append(header_data)
#print(header_list)


final_data = []
for r in range(1,rows+1):
    list_temp = []
    for c in range (1,columns+1):
        row_temp = str(r)
        col_temp = str(c)
        data = driver.find_element(By.XPATH,"/html/body/div[1]/div[2]/section/div[2]/div/div/div[2]/table/tbody/tr["+row_temp+"]/td["+col_temp+"]").text
        print(data, '     ', end ='' )
        list_temp.append(data)
    print()
    #print('{0:<25} {1:>10} {2:>4} {3:>4} {4:>30} {5:>30}'.format(*list_temp))
    final_data.append(list_temp)

#/html/body/div[1]/div[2]/section/div[2]/div/div/div[2]/table/tbody/tr[1]
#/html/body/div[1]/div[2]/section/div[2]/div/div/div[2]/table/tbody/tr[1]/td[1]

patterns = [
    ('Date', lambda text: style(text, fg='white')),
    ('Amount', lambda text: style(text, fg='cyan')),
    ('Balance', lambda text: style(text, fg='green')),
    ('Units', lambda text: style(text, fg='white')),
    ('Trantype', lambda text: style(text, fg='red')),
    ('Terminal', lambda text: style(text, fg='yellow')),
]

table = columnar(final_data, headers=header_list, patterns=patterns, wrap_max=0,max_column_width=30,terminal_width=200)
print(balance_table)
print(table)


################################Excel#########################################




wb_transactions = Workbook()
ws_transactions = wb_transactions.active
wb_transactions.create_sheet('Watcard Balances')
ws_balances = wb_transactions['Watcard Balances']
ws_transactions.title = 'Watcard Transactions'
ws_transactions.append(header_list)
for transaction in final_data:
    ws_transactions.append(transaction)


ws_balances.append(balance_header_list)
for balance_type in balance_final_data:
    ws_balances.append(balance_type)

wb_transactions.save('WatcardTransactions.xlsx')

###########################################################################################

####################################END PROGRAM#########################################################################


#An implicit wait tells WebDriver to poll the DOM for a certain amount of time
# when trying to find any element (or elements)
# not immediately available.
# The default setting is 0 (zero).
# Once set, the implicit wait is set for the life of the WebDriver object. ????
#to poll the DOM for a certain amount of time means means to check the DOM repeatedly, on a set interval (every X milliseconds), to see if an element exists.

#Explicit wait means to wait until the element appears

#driver.back()  can go back to the last page
